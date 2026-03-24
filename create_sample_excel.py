import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Machines"

# Add headers
headers = ['id', 'name', 'status', 'risk', 'load', 'temp', 'vibration', 'pressure', 'rpm', 'health', 'lastService', 'nextService']
ws.append(headers)

# Style the header row
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add sample data
sample_data = [
    ['M001', 'Compressor A', 'Running', 12, 78, 65, 0.45, 110, 2400, 85, '2024-01-15', '2024-06-15'],
    ['M002', 'Pump B', 'Warning', 28, 92, 75, 0.72, 95, 1800, 62, '2024-02-10', '2024-05-20'],
    ['M003', 'Motor C', 'Running', 8, 64, 58, 0.32, 105, 1500, 92, '2023-12-20', '2024-07-20'],
    ['M004', 'Turbine D', 'Critical', 65, 88, 92, 1.85, 125, 3200, 28, '2024-01-05', '2024-03-05'],
    ['M005', 'Generator E', 'Running', 15, 71, 62, 0.51, 100, 1200, 80, '2024-02-01', '2024-08-01'],
    ['M006', 'Blower F', 'Warning', 42, 85, 82, 1.12, 115, 2100, 45, '2023-11-10', '2024-04-10'],
]

for row in sample_data:
    ws.append(row)

# Adjust column widths
column_widths = [10, 15, 12, 8, 8, 8, 12, 12, 8, 8, 15, 15]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64 + i)].width = width

# Save the file
file_path = '/Users/lavanyadhoke/Desktop/predective ai maintainence/sample_machines.xlsx'
wb.save(file_path)
print(f"✓ Excel file created: {file_path}")
